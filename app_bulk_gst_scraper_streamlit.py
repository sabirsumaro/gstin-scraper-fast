import streamlit as st
import openpyxl
import tempfile
import time
import pandas as pd
from io import BytesIO
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# Function to create sample Excel as bytes
def to_excel_bytes(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()

# Actual scraping logic
def run_scraper(input_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    result_wb = openpyxl.Workbook()
    result_ws = result_wb.active
    result_ws.title = "Results"
    result_ws.append(["GSTIN", "Trade Name", "Legal Name", "Principal Place", "Additional Place", "Jurisdiction", "Status"])

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    def close_popups():
        try:
            driver.execute_script("""
                let popups = document.querySelectorAll('.popup, .chatbox, .chat-widget, iframe, .intercom-lightweight-app');
                popups.forEach(p => p.style.display='none');
            """)
        except:
            pass

    def get_data(label):
        try:
            element = driver.find_element(By.XPATH, f"//strong[contains(text(), '{label}')]")
            return element.find_element(By.XPATH, "..").text.replace(label, "").strip()
        except:
            return ""

    for row in ws.iter_rows(min_row=3, min_col=1, max_col=1):
        gstin = row[0].value
        if not gstin:
            continue
        try:
            driver.get("https://irisgst.com/irisperidot/")
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, "gstinno")))
            close_popups()
            input_box = driver.find_element(By.ID, "gstinno")
            input_box.clear()
            input_box.send_keys(gstin)
            driver.find_element(By.XPATH, "//button[contains(text(), 'SEARCH')]").click()

            start_time = time.time()
            while time.time() - start_time < 20:
                if "Trade Name -" in driver.page_source:
                    break
                time.sleep(1)

            close_popups()

            result_ws.append([
                gstin,
                get_data("Trade Name -"),
                get_data("Legal Name of Business -"),
                get_data("Principal Place of Business -"),
                get_data("Additional Place of Business -"),
                get_data("State Jurisdiction -"),
                "Success"
            ])
        except Exception as e:
            result_ws.append([gstin, "", "", "", "", "", f"Error: {str(e)}"])

    driver.quit()
    output_path = input_path.replace(".xlsx", "_Result.xlsx")
    result_wb.save(output_path)
    return output_path

# Streamlit App UI
st.set_page_config(page_title="GSTIN Scraper", layout="centered")
st.title("GSTIN Bulk Scraper 🔍")
st.markdown("Upload your Excel file with GSTINs and get back the extracted results.")

# Sample Template Download
sample_data = pd.DataFrame({'GSTIN': ['06ABCDE1234F1Z5', '07XYZAB1234L1Z2']})
with st.expander("📥 Download Sample Template"):
    excel_bytes = to_excel_bytes(sample_data)
    st.download_button("📥 Download Excel Template", data=excel_bytes, file_name="Bulk_GSTIN_Input_Template.xlsx")

# File Upload + Processing
uploaded_file = st.file_uploader("📤 Upload your GSTIN Excel file", type=["xlsx"])
if uploaded_file is not None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

    with st.spinner("⏳ Processing GSTINs... Please wait..."):
        output_file = run_scraper(tmp_path)

    with open(output_file, "rb") as f:
        st.success("✅ Done! Click below to download your result:")
        st.download_button("📄 Download Result Excel", f, file_name="Bulk_GSTIN_Result.xlsx")
